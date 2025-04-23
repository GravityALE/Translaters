import paramiko
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import simpledialog, messagebox

# Function to connect to the remote device and execute the 'show run' command
def get_service_instances(hostname, port, username, password):
    service_instances = []
    current_interface = None  # Variable to store the current interface
    try:
        # Connect to the remote device using Paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname, port=port, username=username, password=password)

        # Execute the 'show run' command
        stdin, stdout, stderr = ssh.exec_command('show run')
        output = stdout.read().decode()

        # Extract blocks that begin with 'service instance' and end with '!'
        service_block = []
        in_block = False

        for line in output.splitlines():
            if line.strip().startswith('interface'):  # Detect the interface
                current_interface = line.strip()  # Store the interface name
            if line.strip().startswith('service instance'):
                in_block = True
                service_block = [line]  # Start a new service block
            elif in_block:
                service_block.append(line)
                if line.strip() == '!':
                    service_instances.append((current_interface, service_block))  # Add interface + block
                    in_block = False
                    service_block = []

        ssh.close()
    except Exception as e:
        messagebox.showerror("Connection Error", f"An error occurred: {e}")
    return service_instances

# Function to extract key values from a configuration block
def extract_values(service_block):
    service_instance = ""
    description = ""
    encapsulation = ""
    mtu = None
    rewrite = False
    xconnect_ip = ""
    xconnect_id = ""
    backup_peer_ip = ""
    backup_peer_id = ""
    bridge_domain_found = False  # Variable para identificar si se encontró bridge-domain
    counter = 1


    for line in service_block:
        if line.strip().startswith("service instance"):
            service_instance = line.split(" ")[3]
        elif line.strip().startswith("description"):
            description = line.replace("description ", "").strip()
        elif line.strip().startswith("encapsulation dot1q"):
            encapsulation = line.split(" ")[-1]
        elif line.strip().startswith("mtu"):
            mtu = int(line.split(" ")[-1])
            if mtu == 9180:
                if rewrite:
                    mtu = 9198
                else:
                    mtu = 9194
            elif mtu == 9202:
                mtu = None
        elif line.strip().startswith("rewrite ingress tag pop 1 symmetric"):
            rewrite = True
        elif line.strip().startswith("xconnect"):
            parts = line.split(" ")
            xconnect_ip = parts[3]
            xconnect_id = parts[4]
        elif line.strip().startswith("backup peer"):
            parts = line.split(" ")
            backup_peer_ip = parts[5]
            backup_peer_id = parts[6]
        elif line.strip().startswith("bridge-domain"):  # Detectar bridge-domain
            bridge_domain_found = True

    return service_instance, description, encapsulation, mtu, rewrite, xconnect_ip, xconnect_id, backup_peer_ip, backup_peer_id, bridge_domain_found

# Definir los colores de fondo
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro

# En la función save_to_excel_with_template, antes de agregar los valores de la primera fila
def save_to_excel_with_template(service_instances, excel_filename, hostname):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Service Instances"

    # Ajustar el ancho de las columnas
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 3
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 3
    sheet.column_dimensions['E'].width = 50

    # Definir la primera fila con los valores y colores de fondo
    sheet.cell(row=1, column=1, value=f"Configuración ASR {hostname}")
    sheet.cell(row=1, column=1).fill = light_blue_fill  # Aplicar color azul claro a la columna A

    sheet.cell(row=1, column=3, value=f"Subinterfaz {hostname}")
    sheet.cell(row=1, column=3).fill = light_green_fill  # Aplicar color verde claro a la columna C

    sheet.cell(row=1, column=5, value=f"L2VPN {hostname}")
    sheet.cell(row=1, column=5).fill = light_green_fill  # Aplicar color verde claro a la columna E

    # Definir un estilo de relleno para resaltar (fondo amarillo)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    counter = 1
    row = 2
    for current_interface, instance in service_instances:
        service_instance, description, encapsulation, mtu, rewrite, xconnect_ip, xconnect_id, backup_peer_ip, backup_peer_id, bridge_domain_found = extract_values(instance)

        # Incluir la interfaz en la salida y aplicar resalte
        interface_cell = sheet.cell(row=row, column=1, value=f"Interface: {current_interface}")
        interface_cell.fill = highlight_fill  # Aplicar el resaltado a la celda de la interfaz
        row += 1

        part1 = f"""interface Ten0/0/0/#.{service_instance} l2transport
  description {description}
  encapsulation dot1q {encapsulation}"""
        if rewrite:
            part1 += "\n  rewrite ingress tag pop 1 symmetric"

        if mtu:
            part1 += f"\n  mtu {mtu}"



        # Aquí se determina qué parte 2 se utilizará
        if bridge_domain_found:
            part2 = f""" bridge group VPLS_SERVICES
    bridge-domain {encapsulation}
    mtu 9180
interface Ten0/0/0/#.{service_instance}
pw-class MPLS_DEFAULT_ARP"""
        else:
            part2 = f"""xconnect group L2VPN_SERVICES
  p2p L2VPN-{service_instance}-#-{counter}
interface Ten0/0/0/#.{service_instance}
   neighbor ipv4 {xconnect_ip} pw-id {xconnect_id}
pw-class MPLS_DEFAULT_ARP"""
        counter += 1

        if backup_peer_ip and backup_peer_id:
            part2 += f"\n  backup neighbor {backup_peer_ip} pw-id {backup_peer_id}\n  pw-class MPLS_DEFAULT_ARP"

        part2 += "\n   !"

        start_row = row
        for line in instance:
            sheet.cell(row=row, column=1, value=line)
            row += 1

        part1_lines = part1.split('\n')
        part2_lines = part2.split('\n')

        max_lines = max(len(part1_lines), len(part2_lines))

        for i in range(max_lines):
            if i < len(part1_lines):
                sheet.cell(start_row + i, 3, value=part1_lines[i])
            if i < len(part2_lines):
                sheet.cell(start_row + i, 5, value=part2_lines[i])

        row = start_row + 12

    workbook.save(excel_filename)

# Function to handle the user input and run the main process
def run_application():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Get user input
    hostname = simpledialog.askstring("Input", "Ingresa la direccion IP del equipo:")
    username = simpledialog.askstring("Input", "Ingresa el Usuario:")
    password = simpledialog.askstring("Input", "Ingresa la contraseña:", show='*')

    if hostname and username and password:
        service_instances = get_service_instances(hostname, 22, username, password)
        save_to_excel_with_template(service_instances, 'Servicios_ConfiguradosXE_to_XR.xlsx', hostname)
        messagebox.showinfo("Success", "Los datos han sido guardados exitosamente en Servicios_Configurados.xlsx")

# Run the application
if __name__ == "__main__":
    run_application()
