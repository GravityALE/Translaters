import time
import re
from tkinter import messagebox
import paramiko
from openpyxl import Workbook
import tkinter as tk
from tkinter import simpledialog
vsi_counter = 1
# Crear la ventana principal de tkinter (aunque no se mostrará)
root = tk.Tk()
root.withdraw()  # Oculta la ventana principal

# Solicitar credenciales mediante cuadros de diálogo
hostname = simpledialog.askstring("Equipo", "Ingrese el hostname o IP del equipo:", parent=root)
username = simpledialog.askstring("Usuario", "Ingrese el nombre de usuario:", parent=root)
password = simpledialog.askstring("Contraseña", "Ingrese la contraseña:", show="*", parent=root)

# Comando inicial para obtener la descripción de interfaces
command = "show ipv4 interface brief"

# Expresión regular para detectar subinterfaces
pattern = r'^(Gi|Te|Hu|Fa|XGi|Bu|Eth|Po|Vl|Lo)\S+\.\d+'

# Crear un nuevo libro de Excel
wb = Workbook()
ws = wb.active
ws.title = "Resultados"
ws.append(["Configuración de Interfaz", " ", "Información L2VPN", " ", "Plantilla Configuración"])  # Encabezados
# Ajustar el ancho de las columnas A, B, C, D y E
ws.column_dimensions['A'].width = 50  # Ancho de la columna A
ws.column_dimensions['B'].width = 3  # Ancho de la columna B
ws.column_dimensions['C'].width = 50  # Ancho de la columna C
ws.column_dimensions['D'].width = 3  # Ancho de la columna D
ws.column_dimensions['E'].width = 50  # Ancho de la columna E
ws.column_dimensions['F'].width = 3  # Ancho de la columna F
ws.column_dimensions['G'].width = 50
# Función para ejecutar comandos con reconexión si la sesión se cierra, con retraso entre comandos

def execute_command_with_reconnect(client, command, max_retries=5, delay_between_commands=2):
    retries = 0
    while retries < max_retries:
        try:
            stdin, stdout, stderr = client.exec_command(command)
            time.sleep(delay_between_commands)  # Añade un retraso después de ejecutar el comando
            return stdout.read().decode()
        except (paramiko.SSHException, EOFError, OSError):
            print("Error en la conexión. Intentando reconectar...")
            client.close()
            time.sleep(2)  # Espera antes de reintentar
            try:
                client.connect(hostname, username=username, password=password)
                print("Reconexión exitosa.")
            except Exception as e:
                print(f"Error al reconectar: {e}")
                retries += 1
    raise Exception("Error al ejecutar el comando después de múltiples intentos")


try:
    # Crear un cliente SSH
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname, username=username, password=password)

    # Ejecutar el comando inicial para obtener las interfaces
    output = execute_command_with_reconnect(client, command)

    # Filtrar e identificar subinterfaces
    subinterfaces = []
    for line in output.splitlines():
        match = re.match(pattern, line.strip())
        if match:
            subinterfaces.append(match.group())
    print("Subinterfaces encontradas:")
    for subinterface in subinterfaces:
        print(subinterface)

        # Ejecutar el primer comando para cada subinterface
        command_details = f"show run interface {subinterface}"
        details_output = execute_command_with_reconnect(client, command_details).splitlines()

        # Variables para almacenar datos identificados
        description = ""
        mtu = ""
        encapsulation = ""
        subinterface_number = ""
        has_rewrite_tag = False  # Variable para detectar 'rewrite ingress tag pop 1 symmetric'

        # Extraer los valores específicos del primer comando
        for line in details_output:
            if line.strip().startswith("description"):
                description = line.strip().split(" ", 1)[1]  # Extrae la descripción
            elif line.strip().startswith("mtu"):
                mtu = line.strip().split(" ")[1]  # Extrae el valor MTU
            elif line.strip().startswith("encapsulation dot1q"):
                encapsulation = line.strip().split(" ")[-1]  # Extrae el número de encapsulación
            elif line.strip().startswith("rewrite ingress tag pop 1 symmetric"):
                has_rewrite_tag = True  # Marca si existe la línea de rewrite tag
            elif line.strip().startswith("interface"):
                # Extrae el número de subinterfaz desde el nombre de la interfaz
                interface_match = re.search(r"\.(\d+)", line.strip())
                if interface_match:
                    subinterface_number = interface_match.group(1)

        # Ejecutar el segundo comando "show run formal l2vpn | inc <subinterface>"
        command_l2vpn = f"show run formal l2vpn | include interface {subinterface}$"
        l2vpn_output = execute_command_with_reconnect(client, command_l2vpn).splitlines()

        # Procesar todas las líneas de salida L2VPN
        found_command = False
        ipv4 = ""
        pw_id = ""
        backup_ipv4 = ""
        backup_pw_id = ""

        for line in l2vpn_output:
            l2vpn_pattern = r"l2vpn xconnect group (\S+) p2p (\S+) interface (\S+)"
            match = re.match(l2vpn_pattern, line)

            if match:
                group_name, p2p_name, interface = match.groups()
                third_command = f"show run l2vpn xconnect group {group_name} p2p {p2p_name}"
                
                # Ejecutar el tercer comando
                third_output = execute_command_with_reconnect(client, third_command).splitlines()

                # Extraer ipv4, pw-id y datos adicionales del resultado del tercer comando
                for line in third_output:
                    ipv4_pw_pattern = r"neighbor ipv4 (\S+) pw-id (\d+)"
                    ipv4_pw_match = re.search(ipv4_pw_pattern, line)
                    if ipv4_pw_match and not ipv4:
                        ipv4 = ipv4_pw_match.group(1)
                        pw_id = ipv4_pw_match.group(2)
                    elif "backup neighbor" in line:
                        backup_match = re.search(r"backup neighbor (\S+) pw-id (\d+)", line)
                        if backup_match:
                            backup_ipv4 = backup_match.group(1)
                            backup_pw_id = backup_match.group(2)

                # Guardar resultados en Excel
                row_start = ws.max_row + 1
                for line in details_output:
                    ws.append([line, " ", ""])

                # Añadir las líneas del tercer comando en la columna C
                row = row_start
                for line in third_output:
                    ws.cell(row=row, column=3, value=line)
                    row += 1

                 # Generar plantillas según escenario
                mtu_value = "9180" if mtu in ["9194", "9198"] else (mtu if mtu else "9202")
                
                # CASO 1: Interfaces SIN backup (no-VSI)
                if not backup_ipv4:
                    plantilla = f"""
interface GigabitEthernet0/%/#.{subinterface_number} mode l2
 mtu {mtu_value}
 description {description}
 encapsulation dot1q vid {encapsulation}
"""
                    if has_rewrite_tag:
                        plantilla += " rewrite pop single\n"
                    plantilla += f" mpls l2vc {ipv4} {pw_id} control-word raw\n"

                    # Escribir en columna E (para casos no-VSI)
                    row = row_start
                    for line in plantilla.strip().splitlines():
                        ws.cell(row=row, column=5, value=line.strip())
                        row += 1

                # CASO 2: Interfaces CON backup (VSI)
                else:
                    # Plantilla VSI (columna E)
                    plantilla_vsi = f"""
vsi {p2p_name} bd-mode
 pwsignal ldp
  vsi-id {pw_id}
  control-word enable
  flow-label both
  peer {ipv4}
  peer {backup_ipv4} negotiation-vc-id {backup_pw_id}
  protect-group RDCO_{subinterface_number}_{vsi_counter}
   protect-mode pw-redundancy master
   peer {ipv4} preference 1
   peer {backup_ipv4} negotiation-vc-id {backup_pw_id} preference 2
 mtu {mtu_value}
 encapsulation ethernet
"""
                    vsi_counter += 1  # Incrementar el contador para la próxima VSI                   
# Plantilla interfaz + bridge-domain (columna G)
                    plantilla_interfaz = f"""
bridge-domain {encapsulation}
 L2 binding vsi {p2p_name}
#
interface GigabitEthernet0/%/#.{subinterface_number} mode l2
 description {description}
 encapsulation dot1q vid {encapsulation}
 rewrite pop single
 bridge-domain {encapsulation}
#
"""
                    
                    # Escribir ambas plantillas
                    row = row_start
                    for line in plantilla_vsi.strip().splitlines():
                        ws.cell(row=row, column=5, value=line.strip())
                        row += 1
                    
                    row = row_start
                    for line in plantilla_interfaz.strip().splitlines():
                        ws.cell(row=row, column=7, value=line.strip())
                        row += 1


        if not found_command:
            print("No se encontró un formato válido en las líneas del resultado L2VPN.")

finally:
    # Guardar el archivo de Excel
    wb.save("resultados_interfaces_l2vpn.xlsx")
    client.close()
    
    # Mostrar mensaje de confirmación
messagebox.showinfo("Guardado Exitoso", "Se ha guardado satisfactoriamente los datos en 'resultados_interfaces_l2vpn.xlsx'") 