import paramiko
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import simpledialog, messagebox

# Function to connect to the remote device and execute the 'show run' command
def get_service_instances(hostname, port, username, password):
    service_instances = []
    current_interface = None  # Variable to store the current interface
    try:
        # Connect to the remote device using Paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname, port=port, username=username, password=password)

        # Execute the 'show run' command
        stdin, stdout, stderr = ssh.exec_command('show run')
        output = stdout.read().decode()

        # Extract blocks that begin with 'service instance' and end with '!'
        service_block = []
        in_block = False

        for line in output.splitlines():
            if line.strip().startswith('interface'):  # Detect the interface
                current_interface = line.strip()  # Store the interface name
            if line.strip().startswith('service instance'):
                in_block = True
                service_block = [line]  # Start a new service block
            elif in_block:
                service_block.append(line)
                if line.strip() == '!':
                    service_instances.append((current_interface, service_block))  # Add interface + block
                    in_block = False
                    service_block = []

        ssh.close()
    except Exception as e:
        messagebox.showerror("Connection Error", f"An error occurred: {e}")
    return service_instances

# Function to extract key values from a configuration block
def extract_values(service_block):
    service_instance = ""
    description = ""
    encapsulation = ""
    mtu = None
    rewrite = False
    xconnect_ip = ""
    xconnect_id = ""
    backup_peer_ip = ""
    backup_peer_id = ""

    for line in service_block:
        if line.strip().startswith("service instance"):
            service_instance = line.split(" ")[3]
        elif line.strip().startswith("description"):
            description = line.replace("description ", "").strip()
        elif line.strip().startswith("encapsulation dot1q"):
            encapsulation = line.split(" ")[-1]
        elif line.strip().startswith("mtu"):
            mtu = int(line.split(" ")[-1])
            # Apply conditions to mtu
            if mtu == 9180:
                mtu = 9180
            elif mtu == 9202:
                mtu = 9202
        if  mtu is None:  # If no mtu is specified
                mtu = 9216
        elif line.strip().startswith("rewrite ingress tag pop 1 symmetric"):
            rewrite = True
        elif line.strip().startswith("xconnect"):
            parts = line.split(" ")
            xconnect_ip = parts[3]
            xconnect_id = parts[4]
        elif line.strip().startswith("backup peer"):
            parts = line.split(" ")
            backup_peer_ip = parts[5]
            backup_peer_id = parts[6]

    return service_instance, description, encapsulation, mtu, rewrite, xconnect_ip, xconnect_id, backup_peer_ip, backup_peer_id

# Function to save results to Excel
def save_to_excel_with_template(service_instances, excel_filename, hostname):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Service Instances"

    # Set column widths
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 3
    sheet.column_dimensions['C'].width = 50

    # Define first row with values and background colors
    sheet.cell(row=1, column=1, value=f"Configuración ASR {hostname}")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    sheet.cell(row=1, column=1).fill = light_blue_fill

    sheet.cell(row=1, column=3, value=f"Configuracion a Huawei {hostname}")
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    sheet.cell(row=1, column=3).fill = light_green_fill

    # Highlight style for rows
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row = 2
    for current_interface, instance in service_instances:
        service_instance, description, encapsulation, mtu, rewrite, xconnect_ip, xconnect_id, backup_peer_ip, backup_peer_id = extract_values(instance)

        # Highlight interface
        interface_cell = sheet.cell(row=row, column=1, value=f"Interface: {current_interface}")
        interface_cell.fill = highlight_fill  # Apply highlight to interface cell
        row += 1

        part1 = f"interface GigabitEthernet0/#/$.{service_instance} mode L2"
        part1 += f"\n  mtu {mtu}"
        part1 += f"\n  description {description}"
        part1 += f"\n  encapsulation dot1q vid {encapsulation}"

        if rewrite:
            part1 += "\n  rewrite pop single"

        part1 += f"\n  mpls L2vc {xconnect_ip} {xconnect_id} control-word raw"

        if backup_peer_ip and backup_peer_id:
            part1 += f"\n  mpls l2vc {backup_peer_ip} {backup_peer_id} control-word secondary raw"

        part1 += "\n   !"

        start_row = row
        for line in instance:
            sheet.cell(row=row, column=1, value=line)
            row += 1

        part1_lines = part1.split('\n')

        for i, line in enumerate(part1_lines):
            sheet.cell(start_row + i, 3, value=line)

        row = start_row + len(part1_lines) + 2  # Add 2 extra lines for spacing

    workbook.save(excel_filename)

# Function to handle user input and run the process
def run_application():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Get user input
    hostname = simpledialog.askstring("Input", "Ingresa la direccion IP del equipo:")
    username = simpledialog.askstring("Input", "Ingresa el Usuario:")
    password = simpledialog.askstring("Input", "Ingresa la contraseña:", show='*')

    if hostname and username and password:
        service_instances = get_service_instances(hostname, 22, username, password)
        save_to_excel_with_template(service_instances, 'Servicios_ConfiguradosXE_to_VPR.xlsx', hostname)
        messagebox.showinfo("Success", "Los datos han sido guardados exitosamente en Servicios_Configurados.xlsx")

# Run the application
if __name__ == "__main__":
    run_application()